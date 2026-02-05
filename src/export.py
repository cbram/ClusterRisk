"""
Export Module
Exportiert Risiko-Daten nach Excel/LibreOffice
"""

import pandas as pd
from io import BytesIO
from typing import Dict
from datetime import datetime


def export_to_calc(risk_data: Dict, format: str = 'xlsx') -> bytes:
    """
    Exportiert Risiko-Daten nach Excel oder LibreOffice
    
    Args:
        risk_data: Risiko-Daten Dict
        format: 'xlsx' oder 'ods'
    
    Returns:
        Bytes der exportierten Datei
    """
    
    if format == 'xlsx':
        return _export_to_xlsx(risk_data)
    elif format == 'ods':
        return _export_to_ods(risk_data)
    else:
        raise ValueError(f"Unbekanntes Format: {format}")


def _export_to_xlsx(risk_data: Dict) -> bytes:
    """
    Exportiert nach Excel (.xlsx)
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Übersichts-Sheet
        _create_overview_sheet(risk_data, writer)
        
        # Sheet pro Kategorie
        for category, df in risk_data.items():
            if category == 'total_value':
                continue
            
            sheet_names = {
                'asset_class': 'Anlageklasse',
                'sector': 'Branche_Sektor',
                'currency': 'Währung',
                'positions': 'Einzelpositionen'
            }
            
            sheet_name = sheet_names.get(category, category)
            
            if isinstance(df, pd.DataFrame):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatierung
                worksheet = writer.sheets[sheet_name]
                _format_worksheet(worksheet, df)
    
    output.seek(0)
    return output.getvalue()


def _export_to_ods(risk_data: Dict) -> bytes:
    """
    Exportiert nach LibreOffice (.ods)
    """
    output = BytesIO()
    
    # ODS Export mit pandas
    with pd.ExcelWriter(output, engine='odf') as writer:
        
        # Übersichts-Sheet
        _create_overview_sheet(risk_data, writer)
        
        # Sheet pro Kategorie
        for category, df in risk_data.items():
            if category == 'total_value':
                continue
            
            sheet_names = {
                'asset_class': 'Anlageklasse',
                'sector': 'Branche_Sektor',
                'currency': 'Währung',
                'positions': 'Einzelpositionen'
            }
            
            sheet_name = sheet_names.get(category, category)
            
            if isinstance(df, pd.DataFrame):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output.seek(0)
    return output.getvalue()


def _create_overview_sheet(risk_data: Dict, writer):
    """
    Erstellt ein Übersichts-Sheet mit Zusammenfassung
    """
    overview_data = []
    
    # Metadaten
    overview_data.append(['Portfolio Klumpenrisiko Analyse', ''])
    overview_data.append(['Erstellt am', datetime.now().strftime('%d.%m.%Y %H:%M')])
    overview_data.append(['Gesamt-Portfolio-Wert', f"€ {risk_data['total_value']:,.2f}"])
    overview_data.append(['', ''])
    
    # Zusammenfassung pro Kategorie
    overview_data.append(['Kategorie', 'Anzahl Positionen', 'Größte Position (%)', 'Top-5 Konzentration (%)'])
    
    for category in ['asset_class', 'sector', 'currency', 'positions']:
        df = risk_data[category]
        
        if not df.empty:
            category_names = {
                'asset_class': 'Anlageklasse',
                'sector': 'Branche/Sektor',
                'currency': 'Währung',
                'positions': 'Einzelpositionen'
            }
            
            overview_data.append([
                category_names[category],
                len(df),
                f"{df.iloc[0]['Anteil (%)']:.2f}%",
                f"{df.head(5)['Anteil (%)'].sum():.2f}%"
            ])
    
    overview_df = pd.DataFrame(overview_data)
    overview_df.to_excel(writer, sheet_name='Übersicht', index=False, header=False)


def _format_worksheet(worksheet, df: pd.DataFrame):
    """
    Formatiert ein Excel-Worksheet (nur für .xlsx)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Header formatieren
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Spaltenbreite anpassen
    for col in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col)
        max_length = max(
            len(str(df.columns[col-1])),
            df.iloc[:, col-1].astype(str).str.len().max()
        )
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Prozent-Spalten formatieren
    if 'Anteil (%)' in df.columns:
        pct_col = df.columns.get_loc('Anteil (%)') + 1
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=pct_col)
            
            # Farbe basierend auf Risiko
            if cell.value and float(str(cell.value).replace('%', '').replace(',', '.')) > 10:
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif cell.value and float(str(cell.value).replace('%', '').replace(',', '.')) > 5:
                cell.fill = PatternFill(start_color='FFF9CC', end_color='FFF9CC', fill_type='solid')
